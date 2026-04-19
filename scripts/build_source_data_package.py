#!/usr/bin/env python3
from __future__ import annotations

import csv
import shutil
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "submission_assets" / "source_data"
CSV_DIR = OUT_DIR / "csv"
WORKBOOK_PATH = OUT_DIR / "Source Data.xlsx"
README_PATH = OUT_DIR / "README.md"
ZIP_PATH = OUT_DIR / "Source_Data_package.zip"


@dataclass(frozen=True)
class SourceItem:
    sheet: str
    label: str
    description: str
    path: Path


SOURCE_ITEMS = [
    SourceItem(
        sheet="Fig1_abc",
        label="Figure 1a-c district metrics",
        description=(
            "District-level values used for the Figure 1 maps of nitrogen surplus, calorie "
            "production, and total water demand."
        ),
        path=ROOT
        / "_audit"
        / "Nitrogen-Surplus-restructuring"
        / "outputs"
        / "generated"
        / "figure1"
        / "figure1_panel_abc_joined.csv",
    ),
    SourceItem(
        sheet="Fig1d_state_area",
        label="Figure 1d state crop areas",
        description="State-level cereal areas underlying the Figure 1d crop-distribution panel.",
        path=ROOT
        / "_audit"
        / "Nitrogen-Surplus-restructuring"
        / "outputs"
        / "generated"
        / "figure1"
        / "figure1_panel_d_state_area.csv",
    ),
    SourceItem(
        sheet="Fig2a_pareto",
        label="Figure 2a Pareto frontier points",
        description=(
            "Combined rabi+kharif Pareto points for the realized-price benchmark used in Figure 2a."
        ),
        path=ROOT
        / "data"
        / "generated"
        / "Figure2_equivalent"
        / "Figure2_equivalent_panel_a_combined_by_alpha.csv",
    ),
    SourceItem(
        sheet="Fig2b_values",
        label="Figure 2b deterministic endpoint values",
        description=(
            "National aggregate percentage changes for the water-based and nitrogen-based endpoint "
            "strategies shown as colored bars in the realized-price Figure 2b."
        ),
        path=ROOT
        / "data"
        / "generated"
        / "Figure2_equivalent"
        / "Figure2_equivalent_panel_b_values.csv",
    ),
    SourceItem(
        sheet="Fig2b_whiskers",
        label="Figure 2b whisker summary",
        description=(
            "Bootstrap-derived 95% intervals for each displayed Figure 2b metric under the two "
            "endpoint strategies."
        ),
        path=ROOT
        / "data"
        / "generated"
        / "Figure2_equivalent"
        / "panel_b_bootstrap"
        / "Figure2_equivalent_panel_b_bootstrap_summary.csv",
    ),
    SourceItem(
        sheet="Fig2c_retention",
        label="Figure 2c retention-constraint sweep",
        description=(
            "Combined-system response of nitrogen-surplus reduction to progressively relaxed rice "
            "and wheat retention constraints."
        ),
        path=ROOT
        / "data"
        / "generated"
        / "Figure2_equivalent"
        / "Figure2_equivalent_panel_c_combined.csv",
    ),
    SourceItem(
        sheet="Fig2d_flows",
        label="Figure 2d crop-transition flows",
        description=(
            "Long-format source-to-target crop-area flows underlying the Figure 2d alluvial "
            "summary."
        ),
        path=ROOT
        / "data"
        / "generated"
        / "Figure2_equivalent"
        / "Figure2_equivalent_panel_d_transition_long.csv",
    ),
    SourceItem(
        sheet="Fig2d_areas",
        label="Figure 2d optimized area table",
        description=(
            "District-season crop areas before and after nitrogen-focused optimization used to "
            "construct Figure 2d and downstream state/trade summaries."
        ),
        path=ROOT
        / "data"
        / "generated"
        / "Figure2_equivalent"
        / "Figure2_equivalent_panel_d_optimized_areas.csv",
    ),
    SourceItem(
        sheet="FigS2_seasonal",
        label="Supplementary Figure S2 seasonal Pareto points",
        description=(
            "Season-specific Pareto points for the primary realized-price benchmark shown in "
            "Supplementary Figure S2."
        ),
        path=ROOT
        / "data"
        / "generated"
        / "si_figure2_block"
        / "si_s2_seasonal_pareto_points.csv",
    ),
    SourceItem(
        sheet="FigS3_tradeoffs",
        label="Supplementary Figure S3 seasonal endpoint trade-offs",
        description=(
            "Season-specific percentage changes for the water-focused and nitrogen-focused "
            "endpoints shown in Supplementary Figure S3."
        ),
        path=ROOT
        / "data"
        / "generated"
        / "si_figure2_block"
        / "si_s3_seasonal_tradeoffs.csv",
    ),
    SourceItem(
        sheet="FigS4_retention",
        label="Supplementary Figure S4 seasonal retention sweep",
        description=(
            "Season-specific response of nitrogen-surplus reduction to progressively relaxed "
            "rice and wheat retention constraints shown in Supplementary Figure S4."
        ),
        path=ROOT
        / "data"
        / "generated"
        / "si_figure2_block"
        / "si_s4_cultural_retention.csv",
    ),
    SourceItem(
        sheet="Fig3a_state_area",
        label="Figure 3a displayed state totals",
        description=(
            "State-level original and optimized crop areas for the states displayed in Figure 3a."
        ),
        path=ROOT
        / "data"
        / "generated"
        / "Figure3_equivalent"
        / "Figure3_equivalent_panel_a_display_states.csv",
    ),
    SourceItem(
        sheet="Fig3b_edges",
        label="Figure 3b alternative-cereal trade edges",
        description=(
            "Interstate trade edges for the optimized combined alternative-cereal network shown in "
            "Figure 3b."
        ),
        path=ROOT
        / "data"
        / "generated"
        / "Figure3_equivalent"
        / "Figure3_equivalent_panel_b_alt_trade_edges.csv",
    ),
    SourceItem(
        sheet="Fig3b_nodes",
        label="Figure 3b alternative-cereal node flows",
        description="Node-level inflow and outflow summaries supporting the Figure 3b trade diagram.",
        path=ROOT
        / "data"
        / "generated"
        / "Figure3_equivalent"
        / "Figure3_equivalent_panel_b_alt_node_flows.csv",
    ),
    SourceItem(
        sheet="Fig3c_edges",
        label="Figure 3c rice-wheat trade edges",
        description=(
            "Interstate trade edges for the optimized rice-and-wheat network shown in Figure 3c."
        ),
        path=ROOT
        / "data"
        / "generated"
        / "Figure3_equivalent"
        / "Figure3_equivalent_panel_c_rw_trade_edges.csv",
    ),
    SourceItem(
        sheet="Fig3c_nodes",
        label="Figure 3c rice-wheat node flows",
        description="Node-level inflow and outflow summaries supporting the Figure 3c trade diagram.",
        path=ROOT
        / "data"
        / "generated"
        / "Figure3_equivalent"
        / "Figure3_equivalent_panel_c_rw_node_flows.csv",
    ),
    SourceItem(
        sheet="FigS16a_ratio",
        label="Supplementary Figure S16a realized price / MSP ratios",
        description=(
            "All-India realized-price to MSP ratios for the six focal cereals across 2013-14 to "
            "2017-18."
        ),
        path=ROOT / "data" / "generated" / "all_india_unit_price_to_msp_ratio_2013_14_to_2017_18.csv",
    ),
    SourceItem(
        sheet="FigS16b_trade",
        label="Supplementary Figure S16b terms of trade summary",
        description=(
            "Indicative realized-price terms of trade for alternative cereals relative to rice and "
            "wheat."
        ),
        path=ROOT / "data" / "generated" / "terms_of_trade_summary_2013_14_to_2017_18.csv",
    ),
    SourceItem(
        sheet="TableS10_prices",
        label="Supplementary Table S10 primary revenue price summary",
        description=(
            "Crop-level summary of the 2017-18 primary revenue-benchmark prices used in the "
            "revised optimization."
        ),
        path=ROOT
        / "data"
        / "generated"
        / "primary_revenue_price_summary"
        / "primary_revenue_price_summary.csv",
    ),
    SourceItem(
        sheet="FigS17_values",
        label="Supplementary Figure S17 endpoint sensitivity",
        description=(
            "Scenario-wise endpoint results for the MSP benchmark and hybrid realized-price "
            "sensitivity runs."
        ),
        path=ROOT
        / "data"
        / "generated"
        / "si_revenue_benchmark_endpoint_sensitivity"
        / "si_revenue_benchmark_endpoint_sensitivity_values.csv",
    ),
    SourceItem(
        sheet="FigS18a_pareto_cmp",
        label="Supplementary Figure S18a MSP comparison Pareto points",
        description=(
            "Combined rabi+kharif Pareto points for the district-MSP comparison Figure S18a."
        ),
        path=ROOT / "data" / "generated" / "figure2a_no_historical_cap_core_combined_by_alpha.csv",
    ),
    SourceItem(
        sheet="FigS18b_values_cmp",
        label="Supplementary Figure S18b MSP comparison endpoint values",
        description=(
            "District-MSP comparison endpoint values shown as colored bars in Supplementary Figure S18b."
        ),
        path=ROOT / "data" / "generated" / "figure2b_no_historical_cap_core_values.csv",
    ),
    SourceItem(
        sheet="FigS18b_whisk_cmp",
        label="Supplementary Figure S18b MSP comparison whiskers",
        description=(
            "Bootstrap-derived 95% intervals for the district-MSP comparison Figure S18b."
        ),
        path=ROOT
        / "data"
        / "generated"
        / "figure2b_no_historical_cap_core_whiskers"
        / "figure2b_no_historical_cap_core_whiskers_summary.csv",
    ),
    SourceItem(
        sheet="FigS18c_retent_cmp",
        label="Supplementary Figure S18c MSP comparison retention sweep",
        description=(
            "Combined-system response to retained staple-area constraints for the district-MSP comparison figure."
        ),
        path=ROOT / "data" / "generated" / "figure2c" / "combined_no_historical_caps.csv",
    ),
    SourceItem(
        sheet="FigS18d_flows_cmp",
        label="Supplementary Figure S18d MSP comparison transition flows",
        description=(
            "Long-format source-to-target crop-area flows underlying the district-MSP comparison Figure S18d."
        ),
        path=ROOT
        / "data"
        / "generated"
        / "figure2d_no_historical_cap_core"
        / "figure2d_no_historical_cap_core_transition_long.csv",
    ),
    SourceItem(
        sheet="FigS18d_areas_cmp",
        label="Supplementary Figure S18d MSP comparison optimized areas",
        description=(
            "District-season crop areas used to construct the district-MSP comparison Figure S18d."
        ),
        path=ROOT
        / "data"
        / "generated"
        / "figure2d_no_historical_cap_core"
        / "figure2d_no_historical_cap_core_optimized_areas.csv",
    ),
    SourceItem(
        sheet="FigS19a_state_cmp",
        label="Supplementary Figure S19a MSP comparison state areas",
        description=(
            "State-level original and optimized crop areas for the district-MSP comparison Figure S19a."
        ),
        path=ROOT
        / "data"
        / "generated"
        / "figure3a_state_area_comparison"
        / "figure3a_state_area_comparison_display_states.csv",
    ),
    SourceItem(
        sheet="FigS19b_edges_cmp",
        label="Supplementary Figure S19b MSP comparison alt-cereal edges",
        description=(
            "Alternative-cereal trade edges for the district-MSP comparison Figure S19b."
        ),
        path=ROOT
        / "data"
        / "generated"
        / "figure3_trade_networks"
        / "figure3b_alt_trade_edges_clean.csv",
    ),
    SourceItem(
        sheet="FigS19b_nodes_cmp",
        label="Supplementary Figure S19b MSP comparison alt-cereal nodes",
        description=(
            "Alternative-cereal node-flow summaries for the district-MSP comparison Figure S19b."
        ),
        path=ROOT
        / "data"
        / "generated"
        / "figure3_trade_networks"
        / "figure3b_alt_node_flows_clean.csv",
    ),
    SourceItem(
        sheet="FigS19c_edges_cmp",
        label="Supplementary Figure S19c MSP comparison rice-wheat edges",
        description=(
            "Rice-wheat trade edges for the district-MSP comparison Figure S19c."
        ),
        path=ROOT
        / "data"
        / "generated"
        / "figure3_trade_networks"
        / "figure3c_rice_wheat_trade_edges_clean.csv",
    ),
    SourceItem(
        sheet="FigS19c_nodes_cmp",
        label="Supplementary Figure S19c MSP comparison rice-wheat nodes",
        description=(
            "Rice-wheat node-flow summaries for the district-MSP comparison Figure S19c."
        ),
        path=ROOT
        / "data"
        / "generated"
        / "figure3_trade_networks"
        / "figure3c_rice_wheat_node_flows_clean.csv",
    ),
    SourceItem(
        sheet="FigS20_frontier",
        label="Supplementary Figure S20 realized-price frontier envelope summary",
        description=(
            "Bootstrap summary for the primary realized-price Pareto frontier envelope."
        ),
        path=ROOT
        / "data"
        / "generated"
        / "Figure2_equivalent_frontier_bootstrap"
        / "Figure2_equivalent_frontier_bootstrap_summary.csv",
    ),
    SourceItem(
        sheet="FigS21_sum",
        label="Supplementary Figure S21 primary realized-price seasonal transition summary",
        description=(
            "Top seasonal transition flows used to interpret the seasonally disaggregated "
            "primary realized-price crop-substitution audit."
        ),
        path=ROOT
        / "data"
        / "generated"
        / "seasonal_substitution_audit_primary_revenue"
        / "seasonal_top_non_diagonal_transitions.csv",
    ),
    SourceItem(
        sheet="FigS21_flags",
        label="Supplementary Figure S21 primary realized-price district-season flags",
        description=(
            "District-season audit table identifying rice-loss / wheat-gain or wheat-loss / "
            "rice-gain co-adjustments under the primary realized-price benchmark."
        ),
        path=ROOT
        / "data"
        / "generated"
        / "seasonal_substitution_audit_primary_revenue"
        / "district_season_rice_wheat_flags.csv",
    ),
]


HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
TITLE_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")


def relpath(path: Path) -> str:
    return str(path.relative_to(ROOT))


def ensure_dirs() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    CSV_DIR.mkdir(parents=True, exist_ok=True)


def write_readme(manifest_rows: list[dict[str, object]]) -> None:
    lines = [
        "# Source Data package",
        "",
        "This folder contains the Nature-style source-data package prepared for revision 2 of",
        "\"Quantifying Environmental Co-Benefits of Nitrogen-Based Crop Restructuring and Its",
        "Implications on India's Interstate Trade Network.\"",
        "",
        "Primary artifact:",
        "",
        "- `Source Data.xlsx`: workbook containing figure-ready source data tables and the primary revenue-price summary table.",
        "",
        "Supporting artifacts:",
        "",
        "- `csv/`: CSV mirrors of the workbook sheets.",
        "- `Source_Data_package.zip`: convenience archive of this folder for submission handling.",
        "",
        "Workbook coverage:",
        "",
        "- Main manuscript Figures 1 to 3.",
        "- Supplementary Figures S2 to S4 and S16 to S21 introduced or revised during the current revision round.",
        "- Supplementary Table S10 summarizing the primary realized-price revenue benchmark.",
        "",
        "Sheet manifest:",
        "",
        "| Sheet | Display item | Rows | Columns | CSV mirror | Relative source path |",
        "| --- | --- | ---: | ---: | --- | --- |",
    ]
    for row in manifest_rows:
        lines.append(
            f"| `{row['sheet']}` | {row['label']} | {row['rows']} | {row['columns']} | "
            f"`csv/{row['csv_name']}` | `{row['relative_path']}` |"
        )
    lines.append("")
    lines.append(
        "The broader public input datasets and repository-level reproducibility workflow are described "
        "separately in the manuscript Data Availability and Code Availability statements."
    )
    README_PATH.write_text("\n".join(lines) + "\n")


def autosize_columns(ws) -> None:
    for col_cells in ws.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in col_cells]
        width = min(max(len(v) for v in values) + 2, 48)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width


def add_readme_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "README"
    rows = [
        ["Source Data package", ""],
        ["Manuscript", "Quantifying Environmental Co-Benefits of Nitrogen-Based Crop Restructuring and Its Implications on India's Interstate Trade Network"],
        ["Revision", "Nature Communications revision 2"],
        ["Contents", "Source data workbook for Figs. 1-3, Supplementary Figs. S2-S4 and S16-S21, and Supplementary Table S10"],
        ["Notes", "Each figure sheet contains the table used directly to construct the corresponding display item. See the Manifest sheet for file provenance and row counts."],
    ]
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.fill = TITLE_FILL
            elif c_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = HEADER_FILL
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 110


def add_manifest_sheet(wb: Workbook, manifest_rows: list[dict[str, object]]) -> None:
    ws = wb.create_sheet("Manifest")
    headers = ["Sheet", "Display item", "Description", "Rows", "Columns", "CSV mirror", "Relative source path"]
    ws.append(headers)
    for header_cell in ws[1]:
        header_cell.font = Font(bold=True)
        header_cell.fill = HEADER_FILL
        header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in manifest_rows:
        ws.append(
            [
                row["sheet"],
                row["label"],
                row["description"],
                row["rows"],
                row["columns"],
                f"csv/{row['csv_name']}",
                row["relative_path"],
            ]
        )
    ws.freeze_panes = "A2"
    autosize_columns(ws)


def style_data_sheet(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    autosize_columns(ws)


def copy_csv(item: SourceItem) -> str:
    csv_name = f"{item.sheet}.csv"
    target = CSV_DIR / csv_name
    shutil.copy2(item.path, target)
    return csv_name


def write_data_sheet(wb: Workbook, item: SourceItem, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(item.sheet)
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))
    style_data_sheet(ws)


def build_workbook() -> list[dict[str, object]]:
    ensure_dirs()
    wb = Workbook()
    add_readme_sheet(wb)

    manifest_rows: list[dict[str, object]] = []
    for item in SOURCE_ITEMS:
        if not item.path.exists():
            raise FileNotFoundError(f"Missing source-data file: {item.path}")
        df = pd.read_csv(item.path)
        csv_name = copy_csv(item)
        write_data_sheet(wb, item, df)
        manifest_rows.append(
            {
                "sheet": item.sheet,
                "label": item.label,
                "description": item.description,
                "rows": int(df.shape[0]),
                "columns": int(df.shape[1]),
                "csv_name": csv_name,
                "relative_path": relpath(item.path),
            }
        )

    add_manifest_sheet(wb, manifest_rows)
    wb.save(WORKBOOK_PATH)
    write_readme(manifest_rows)
    if ZIP_PATH.exists():
        ZIP_PATH.unlink()
    shutil.make_archive(str(ZIP_PATH.with_suffix("")), "zip", OUT_DIR)
    return manifest_rows


def main() -> None:
    manifest_rows = build_workbook()
    print(f"workbook: {WORKBOOK_PATH}")
    print(f"zip: {ZIP_PATH}")
    print(f"sheets: {len(manifest_rows) + 2}")
    print("items:")
    for row in manifest_rows:
        print(f"  - {row['sheet']}: {row['rows']} rows x {row['columns']} cols")


if __name__ == "__main__":
    main()
